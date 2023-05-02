import tkinter as tk
from tkinter import filedialog
import openpyxl
from tkcalendar import DateEntry
from tkinter import ttk
from tkinter import messagebox
from openpyxl.formula.translate import Translator

def check_credentials():
    # Check if the entered username and password are correct
    if username_entry.get() == "lacova" and password_entry.get() == "lacova":
        # If they are correct, hide the login window and show the main application window
        password_entry.delete(0, tk.END)
        login_window.withdraw()
        window.deiconify()
    else:
        # If they are incorrect, show an error message as a pop-up
        messagebox.showerror("Error", "Incorrect username or password")
        
        # Clear the input fields
        username_entry.delete(0, tk.END)
        password_entry.delete(0, tk.END)   

def logout():
    # Hide the main application window and show the login window again
    window.withdraw()
    login_window.deiconify()


# Create a login window
login_window = tk.Tk()
login_window.geometry("475x420")

login_window.title("Lacova Login")

logo_image = tk.PhotoImage(file="fb.png")
logo_label = tk.Label(login_window, image=logo_image, bg="#3b5998")
logo_label.place(x=0, y=0)
# Create the labels and input fields for the username and password
username_label = tk.Label(login_window, text="Username", fg="white", bg="#3b5998", font=("Helvetica", 12))
username_label.place(x=50, y=150)
# tk.Label(login_window, text="Username:").grid(row=5, column=0)
username_entry = tk.Entry(login_window, bg="#dfe3ee", font=("Helvetica", 12))
username_entry.place(x=50, y=180)

password_label = tk.Label(login_window, text="Password", fg="white", bg="#3b5998", font=("Helvetica", 12))
password_label.place(x=50, y=220)
password_entry = tk.Entry(login_window, show="*", bg="#dfe3ee", font=("Helvetica", 12))
password_entry.place(x=50, y=250)

submit_button = tk.Button(login_window, text="Log In", command=check_credentials, bg="#1877f2", fg="white", font=("Helvetica", 12), width=25, bd=0)
submit_button.place(x=50, y=300)




# Create a label to display error messages
error_label = tk.Label(login_window, fg="red")
error_label.grid(row=3, column=0, columnspan=2)

# Create the main application window (hidden until login is successful)


def calculate_total():
    # Initialize the total variable to 0
    total = 0

    # Calculate the total for each input field that contains a value
    if input_field4 and input_field3.get():
        total += float(input_field4.get()) * float(input_field3.get())
    if input_field6.get() and input_field5.get():
        total += float(input_field6.get()) * float(input_field5.get())
    if input_field8.get() and input_field7.get():
        total += float(input_field8.get()) * float(input_field7.get())
    if input_field10.get() and input_field9.get():
        total += float(input_field10.get()) * float(input_field9.get())
    if input_field12.get() and input_field11.get():
        total += float(input_field12.get()) * float(input_field11.get())
    if input_field14.get() and input_field13.get():
        total += float(input_field14.get()) * float(input_field13.get())
    if input_field16.get() and input_field15.get():
        total += float(input_field16.get()) * float(input_field15.get())
    if input_field18.get() and input_field17.get():
        total += float(input_field18.get()) * float(input_field17.get())
    if input_field20.get() and input_field19.get():
        total += float(input_field20.get()) * float(input_field19.get())
    if input_field22.get() and input_field21.get():
        total += float(input_field22.get()) * float(input_field21.get())
    if input_field24.get() and input_field23.get():
        total += float(input_field24.get()) * float(input_field23.get())
    if input_field26.get() and input_field25.get():
        total += float(input_field26.get()) * float(input_field25.get())
    if input_field28.get() and input_field27.get():
        total += float(input_field28.get()) * float(input_field27.get())
    if input_field30.get() and input_field29.get():
        total += float(input_field30.get()) * float(input_field29.get())
    if input_field32.get() and input_field31.get():
        total += float(input_field32.get()) * float(input_field31.get())
    if input_field34.get() and input_field33.get():
        total += float(input_field34.get()) * float(input_field33.get())
    if input_field36.get() and input_field35.get():
        total += float(input_field36.get()) * float(input_field35.get())
    if input_field38.get() and input_field37.get():
        total += float(input_field38.get()) * float(input_field37.get())
    if input_field40.get() and input_field39.get():
        total += float(input_field40.get()) * float(input_field39.get())
    if input_field42.get() and input_field41.get():
        total += float(input_field42.get()) * float(input_field41.get())
    # if input_field44.get() and input_field43.get():
    #     total += float(input_field44.get()) * float(input_field43.get())
    if input_field46.get() and input_field45.get():
        total += float(input_field46.get()) * float(input_field45.get())
    if input_field48.get() and input_field47.get():
        total += float(input_field48.get()) * float(input_field47.get())
    if input_field50.get() and input_field49.get():
        total += float(input_field50.get()) * float(input_field49.get())
    
    if input_field52.get() and input_field51.get():
        total += float(input_field52.get()) * float(input_field51.get())
    if input_field54.get() and input_field53.get():
        total += float(input_field54.get()) * float(input_field53.get())

    input_field55.delete(0, tk.END)
    input_field55.insert(0, total)






# Create a tkinter window and add input fields for the data
window = tk.Tk()
window.title("Lacova infinity Global Enterprises")
# Create the labels and input fields
tk.Label(window, text="Enter Date:").grid(row=0, column=0)
input_field = DateEntry(window, width=12, background='darkblue', foreground='white', borderwidth=2)
input_field.grid(row=1, column=0)

tk.Label(window, text="Enter Invoice Number:").grid(row=0, column=1)
input_field1 = tk.Entry(window)
input_field1.grid(row=1, column=1)

tk.Label(window, text="Customer Name:").grid(row=0, column=2)
input_field2 = tk.Entry(window)
input_field2.grid(row=1, column=2)

tk.Label(window, text="QUANTITY of 3.4Kg:").grid(row=0, column=3)
input_field3 = tk.Entry(window)
input_field3.grid(row=1, column=3)

tk.Label(window, text="Price of 3.4Kg:").grid(row=0, column=4)
input_field4 = tk.Entry(window)
input_field4.grid(row=1, column=4)


tk.Label(window, text="QUANTITY of 3.5Kg:").grid(row=0, column=5)
input_field5 = tk.Entry(window)
input_field5.grid(row=1, column=5)

tk.Label(window, text="Price of 3.5Kg:").grid(row=2, column=0)
input_field6 = tk.Entry(window)
input_field6.grid(row=3, column=0)


tk.Label(window, text="QUANTITY of 3.6Kg:").grid(row=2, column=1)
input_field7 = tk.Entry(window)
input_field7.grid(row=3, column=1)

tk.Label(window, text="Price of 3.6Kg:").grid(row=2, column=2)
input_field8 = tk.Entry(window)
input_field8.grid(row=3, column=2)

tk.Label(window, text="QUANTITY of 3.7Kg:").grid(row=2, column=3)
input_field9 = tk.Entry(window)
input_field9.grid(row=3, column=3)

tk.Label(window, text="Price of 3.7Kg:").grid(row=2, column=4)
input_field10 = tk.Entry(window)
input_field10.grid(row=3, column=4)

tk.Label(window, text="QUANTITY of 3.8Kg:").grid(row=2, column=5)
input_field11 = tk.Entry(window)
input_field11.grid(row=3, column=5)

tk.Label(window, text="Price of 3.8Kg:").grid(row=4, column=0)
input_field12 = tk.Entry(window)
input_field12.grid(row=5, column=0)

tk.Label(window, text="QUANTITY of 3.9Kg:").grid(row=4, column=1)
input_field13 = tk.Entry(window)
input_field13.grid(row=5, column=1)

tk.Label(window, text="Price of 3.9Kg:").grid(row=4, column=2)
input_field14 = tk.Entry(window)
input_field14.grid(row=5, column=2)

tk.Label(window, text="QUANTITY of 4.0Kg:").grid(row=4, column=3)
input_field15 = tk.Entry(window)
input_field15.grid(row=5, column=3)

tk.Label(window, text="Price of 4.0Kg:").grid(row=4, column=4)
input_field16 = tk.Entry(window)
input_field16.grid(row=5, column=4)

tk.Label(window, text="QUANTITY of 4.1Kg:").grid(row=4, column=5)
input_field17 = tk.Entry(window)
input_field17.grid(row=5, column=5)

tk.Label(window, text="Price of 4.1Kg:").grid(row=6, column=0)
input_field18 = tk.Entry(window)
input_field18.grid(row=7, column=0)

tk.Label(window, text="QUANTITY of 4.2Kg:").grid(row=6, column=1)
input_field19 = tk.Entry(window)
input_field19.grid(row=7, column=1)

tk.Label(window, text="Price of 4.2Kg:").grid(row=6, column=2)
input_field20 = tk.Entry(window)
input_field20.grid(row=7, column=2)

tk.Label(window, text="QUANTITY of 4.3Kg:").grid(row=6, column=3)
input_field21 = tk.Entry(window)
input_field21.grid(row=7, column=3)

tk.Label(window, text="Price of 4.3Kg:").grid(row=6, column=4)
input_field22 = tk.Entry(window)
input_field22.grid(row=7, column=4)

tk.Label(window, text="QUANTITY of 4.4Kg:").grid(row=6, column=5)
input_field23 = tk.Entry(window)
input_field23.grid(row=7, column=5)

tk.Label(window, text="Price of 4.4Kg:").grid(row=8, column=0)
input_field24 = tk.Entry(window)
input_field24.grid(row=9, column=0)

tk.Label(window, text="QUANTITY of 4.5Kg:").grid(row=8, column=1)
input_field25 = tk.Entry(window)
input_field25.grid(row=9, column=1)

tk.Label(window, text="Price of 4.5Kg:").grid(row=8, column=2)
input_field26 = tk.Entry(window)
input_field26.grid(row=9, column=2)

tk.Label(window, text="QUANTITY of 4.6Kg:").grid(row=8, column=3)
input_field27 = tk.Entry(window)
input_field27.grid(row=9, column=3)

tk.Label(window, text="Price of 4.6Kg:").grid(row=8, column=4)
input_field28 = tk.Entry(window)
input_field28.grid(row=9, column=4)

tk.Label(window, text="QUANTITY of 4.7Kg:").grid(row=8, column=5)
input_field29 = tk.Entry(window)
input_field29.grid(row=9, column=5)

tk.Label(window, text="Price of 4.7Kg:").grid(row=10, column=0)
input_field30 = tk.Entry(window)
input_field30.grid(row=11, column=0)

tk.Label(window, text="QUANTITY of 4.8Kg:").grid(row=10, column=1)
input_field31 = tk.Entry(window)
input_field31.grid(row=11, column=1)

tk.Label(window, text="Price of 4.8Kg:").grid(row=10, column=2)
input_field32 = tk.Entry(window)
input_field32.grid(row=11, column=2)

tk.Label(window, text="QUANTITY of 4.9Kg:").grid(row=10, column=3)
input_field33 = tk.Entry(window)
input_field33.grid(row=11, column=3)

tk.Label(window, text="Price of 4.9Kg:").grid(row=10, column=4)
input_field34 = tk.Entry(window)
input_field34.grid(row=11, column=4)

tk.Label(window, text="QUANTITY of 5.0Kg:").grid(row=10, column=5)
input_field35 = tk.Entry(window)
input_field35.grid(row=11, column=5)

tk.Label(window, text="Price of 5.0Kg:").grid(row=12, column=0)
input_field36 = tk.Entry(window)
input_field36.grid(row=13, column=0)

tk.Label(window, text="QUANTITY of 5.2Kg:").grid(row=12, column=1)
input_field37 = tk.Entry(window)
input_field37.grid(row=13, column=1)

tk.Label(window, text="Price of 5.2Kg:").grid(row=12, column=2)
input_field38 = tk.Entry(window)
input_field38.grid(row=13, column=2)

tk.Label(window, text="QUANTITY of 5.5Kg:").grid(row=12, column=3)
input_field39 = tk.Entry(window)
input_field39.grid(row=13, column=3)

tk.Label(window, text="Price of 5.5Kg:").grid(row=12, column=4)
input_field40 = tk.Entry(window)
input_field40.grid(row=13, column=4)

tk.Label(window, text="QUANTITY of 5.8Kg:").grid(row=12, column=5)
input_field41 = tk.Entry(window)
input_field41.grid(row=13, column=5)

tk.Label(window, text="Price of 5.8Kg:").grid(row=14, column=0)
input_field42 = tk.Entry(window)
input_field42.grid(row=15, column=0)

# tk.Label(window, text="QUANTITY of 5.9Kg:").grid(row=14, column=1)
# input_field43 = tk.Entry(window)
# input_field43.grid(row=15, column=1)

# tk.Label(window, text="Price of 5.9Kg:").grid(row=14, column=2)
# input_field44 = tk.Entry(window)
# input_field44.grid(row=15, column=2)

tk.Label(window, text="QUANTITY of 6.0Kg:").grid(row=14, column=1)
input_field45 = tk.Entry(window)
input_field45.grid(row=15, column=1)

tk.Label(window, text="Price of 6.0Kg:").grid(row=14, column=2)
input_field46 = tk.Entry(window)
input_field46.grid(row=15, column=2)

tk.Label(window, text="QUANTITY of Oloja:").grid(row=14, column=3)
input_field47 = tk.Entry(window)
input_field47.grid(row=15, column=3)

tk.Label(window, text="Price of Oloja:").grid(row=14, column=4)
input_field48 = tk.Entry(window)
input_field48.grid(row=15, column=4)

tk.Label(window, text="QUANTITY of Ice:").grid(row=14, column=5)
input_field49 = tk.Entry(window)
input_field49.grid(row=15, column=5)

tk.Label(window, text="Price of Ice:").grid(row=16, column=0)
input_field50 = tk.Entry(window)
input_field50.grid(row=17, column=0)

tk.Label(window, text="QUANTITY of Dispenser:").grid(row=16, column=1)
input_field51 = tk.Entry(window)
input_field51.grid(row=17, column=1)

tk.Label(window, text="Price of Dispenser:").grid(row=16, column=2)
input_field52 = tk.Entry(window)
input_field52.grid(row=17, column=2)

tk.Label(window, text="QUANTITY of Small:").grid(row=16, column=3)
input_field53 = tk.Entry(window)
input_field53.grid(row=17, column=3)

tk.Label(window, text="Price of Small:").grid(row=16, column=4)
input_field54 = tk.Entry(window)
input_field54.grid(row=17, column=4)


tk.Label(window, text="TOTAL").grid(row=16, column=5)
input_field55 = tk.Entry(window)
input_field55.grid(row=17, column=5)

tk.Label(window, text="Amount Paid").grid(row=18, column=0)
input_field56 = tk.Entry(window)
input_field56.grid(row=19, column=0)

# Create the button to calculate the total
calculate_button = tk.Button(window, text="Calculate Total", command=calculate_total, bg="blue", fg="white", font=("Helvetica", 12), width=10, bd=0)
calculate_button.grid(row=21, column=0)


submitted_label = tk.Label(window, text="")
submitted_label.grid(row=21, column=1, columnspan=3)

not_submitted_label = tk.Label(window, text="")
not_submitted_label.grid(row=21, column=2, columnspan=3)

# Add a button to log out
logout_button = tk.Button(window, text="Log Out", command=logout, bg="red", fg="white", font=("Helvetica", 12), width=10, bd=0)
logout_button.grid(row=23, column=5, columnspan=3)


def save_data():
    if not input_field2.get():
        not_submitted_label.config(text="Fill the neccesary fields")
        messagebox.showinfo("Error", "Input Customer Name")
        # window.after(3000, lambda: messagebox._show("",""))
        not_submitted_label.after(3000, lambda: not_submitted_label.config(text="")) # Remove the text after 5 seconds
        return  # Do nothing if input field is empty
    if not input_field1.get():
        not_submitted_label.config(text="Fill the neccesary fields")
        messagebox.showinfo("Error", "Input invoice Number!")
        not_submitted_label.after(3000, lambda: not_submitted_label.config(text="")) # Remove the text after 5 seconds
        return  # Do nothing if input field is empty
    
    if not input_field56.get():
        not_submitted_label.config(text="Fill the neccesary fields")
        messagebox.showinfo("Error", "Input Amount Paid By Customer")
        not_submitted_label.after(3000, lambda: not_submitted_label.config(text="")) # Remove the text after 5 seconds
        return  # Do nothing if input field is empty

    # Load the existing workbook
    workbook = openpyxl.load_workbook('one.xlsx')

    # Select the sheet you want to modify
    sheet = workbook['Sales Invoice Listing']
    sheet1 = workbook['2023 TURNOVER']
    sheet2 = workbook['Debtors and Receivables']
    sheet3 = workbook['Cash Account Listing']

    # Find the last row in the sheet
    max_row = sheet.max_row
    max_row1 = sheet1.max_row
    max_row2 = sheet2.max_row
    max_row3 = sheet3.max_row
    



    sheet.cell(row=max_row+1, column=1, value=input_field.get())
    sheet.cell(row=max_row+1, column=2, value=input_field1.get())
    sheet.cell(row=max_row+1, column=3, value=input_field2.get())
    sheet.cell(row=max_row+1, column=4, value=input_field3.get())    
    sheet.cell(row=max_row+1, column=5, value=input_field4.get())
    sheet.cell(row=max_row+1, column=6, value=input_field5.get())
    sheet.cell(row=max_row+1, column=7, value=input_field6.get())
    sheet.cell(row=max_row+1, column=8, value=input_field7.get())
    sheet.cell(row=max_row+1, column=9, value=input_field8.get())
    sheet.cell(row=max_row+1, column=10, value=input_field9.get())
    sheet.cell(row=max_row+1, column=11, value=input_field10.get())
    sheet.cell(row=max_row+1, column=12, value=input_field11.get())
    sheet.cell(row=max_row+1, column=13, value=input_field12.get())
    sheet.cell(row=max_row+1, column=14, value=input_field13.get())
    sheet.cell(row=max_row+1, column=15, value=input_field14.get())
    sheet.cell(row=max_row+1, column=16, value=input_field15.get())
    sheet.cell(row=max_row+1, column=17, value=input_field16.get())
    sheet.cell(row=max_row+1, column=18, value=input_field17.get())
    sheet.cell(row=max_row+1, column=19, value=input_field18.get())
    sheet.cell(row=max_row+1, column=20, value=input_field19.get())
    sheet.cell(row=max_row+1, column=21, value=input_field20.get())
    sheet.cell(row=max_row+1, column=22, value=input_field21.get())
    sheet.cell(row=max_row+1, column=23, value=input_field22.get())
    sheet.cell(row=max_row+1, column=24, value=input_field23.get())
    sheet.cell(row=max_row+1, column=25, value=input_field24.get()) 
    sheet.cell(row=max_row+1, column=26, value=input_field25.get())
    sheet.cell(row=max_row+1, column=27, value=input_field26.get())
    sheet.cell(row=max_row+1, column=28, value=input_field27.get())
    sheet.cell(row=max_row+1, column=29, value=input_field28.get())
    sheet.cell(row=max_row+1, column=30, value=input_field29.get())
    sheet.cell(row=max_row+1, column=31, value=input_field30.get())
    sheet.cell(row=max_row+1, column=32, value=input_field31.get())
    sheet.cell(row=max_row+1, column=33, value=input_field32.get())
    sheet.cell(row=max_row+1, column=34, value=input_field33.get())
    sheet.cell(row=max_row+1, column=35, value=input_field34.get())
    sheet.cell(row=max_row+1, column=36, value=input_field35.get())
    sheet.cell(row=max_row+1, column=37, value=input_field36.get())
    sheet.cell(row=max_row+1, column=38, value=input_field37.get())
    sheet.cell(row=max_row+1, column=39, value=input_field38.get())
    sheet.cell(row=max_row+1, column=40, value=input_field39.get())
    sheet.cell(row=max_row+1, column=41, value=input_field40.get())
    sheet.cell(row=max_row+1, column=42, value=input_field41.get())
    sheet.cell(row=max_row+1, column=43, value=input_field42.get())
    sheet.cell(row=max_row+1, column=44, value=input_field45.get())
    sheet.cell(row=max_row+1, column=45, value=input_field46.get())
    sheet.cell(row=max_row+1, column=46, value=input_field47.get())
    sheet.cell(row=max_row+1, column=47, value=input_field48.get())
    sheet.cell(row=max_row+1, column=48, value=input_field49.get())
    sheet.cell(row=max_row+1, column=49, value=input_field50.get())
    sheet.cell(row=max_row+1, column=50, value=input_field51.get())
    sheet.cell(row=max_row+1, column=51, value=input_field52.get())
    sheet.cell(row=max_row+1, column=52, value=input_field53.get())
    sheet.cell(row=max_row+1, column=53, value=input_field54.get())
    sheet.cell(row=max_row+1, column=54, value=input_field55.get())

    sheet3.cell(row=max_row3+1, column=7, value=input_field56.get())

    #LINKING SL TO SA AND RVB
    formula_date_SA = f"='Sales Invoice Listing'!A{max_row + 1}"
    sheet1[f"A{max_row1 + 1}"] = formula_date_SA

    formula_IN_SA = f"='Sales Invoice Listing'!B{max_row + 1}"
    sheet1[f"B{max_row1 + 1}"] = formula_IN_SA


    formula_CN_SA = f"='Sales Invoice Listing'!C{max_row + 1}"
    sheet1[f"C{max_row1 + 1}"] = formula_CN_SA

    formula_TT_SA = f"='Sales Invoice Listing'!BB{max_row + 1}"
    sheet1[f"E{max_row1 + 1}"] = formula_TT_SA


    formula_date_RVB = f"='Sales Invoice Listing'!A{max_row + 1}"
    sheet2[f"A{max_row2 + 1}"] = formula_date_RVB

    formula_IN_RVB = f"='Sales Invoice Listing'!B{max_row + 1}"
    sheet2[f"B{max_row2 + 1}"] = formula_IN_RVB


    formula_CN_RVB = f"='Sales Invoice Listing'!C{max_row + 1}"
    sheet2[f"C{max_row2 + 1}"] = formula_CN_RVB

    formula_TT_RVB = f"='Sales Invoice Listing'!BB{max_row + 1}"
    sheet2[f"D{max_row2 + 1}"] = formula_TT_RVB

    formula_AP_RVB = f"='Cash Account Listing'!G{max_row3 + 1}"
    sheet2[f"E{max_row2 + 1}"] = formula_AP_RVB


    formula_date_CAL = f"='Sales Invoice Listing'!A{max_row + 1}"
    sheet3[f"A{max_row3 + 1}"] = formula_date_CAL

    formula_IN_CAL = f"='Sales Invoice Listing'!B{max_row + 1}"
    sheet3[f"D{max_row3 + 1}"] = formula_IN_CAL


    formula_CN_CAL = f"='Sales Invoice Listing'!C{max_row + 1}"
    sheet3[f"B{max_row3 + 1}"] = formula_CN_CAL

    formula_TT_CAL = f"='Sales Invoice Listing'!BB{max_row + 1}"
    sheet3[f"F{max_row3 + 1}"] = formula_TT_CAL






    # Save the workbook
    workbook.save('one.xlsx')
    # messagebox.showinfo("Success", "Data pushed successfully!")
    # Update the submitted prompt
    submitted_label.config(text="Data pushed successfully")
    submitted_label.after(3000, lambda: submitted_label.config(text="")) # Remove the text after 5 seconds


    input_field.delete(0, tk.END)
    input_field1.delete(0, tk.END)
    input_field2.delete(0, tk.END)
    input_field3.delete(0, tk.END)
    input_field4.delete(0, tk.END)
    input_field5.delete(0, tk.END)
    input_field6.delete(0, tk.END)
    input_field7.delete(0, tk.END)
    input_field8.delete(0, tk.END)
    input_field9.delete(0, tk.END)
    input_field10.delete(0, tk.END)
    input_field11.delete(0, tk.END)
    input_field12.delete(0, tk.END)
    input_field13.delete(0, tk.END)
    input_field14.delete(0, tk.END)
    input_field15.delete(0, tk.END)
    input_field16.delete(0, tk.END)
    input_field17.delete(0, tk.END)
    input_field18.delete(0, tk.END)
    input_field19.delete(0, tk.END)
    input_field20.delete(0, tk.END)
    input_field21.delete(0, tk.END)
    input_field22.delete(0, tk.END)
    input_field23.delete(0, tk.END)
    input_field24.delete(0, tk.END)
    input_field25.delete(0, tk.END)
    input_field26.delete(0, tk.END)
    input_field27.delete(0, tk.END)
    input_field28.delete(0, tk.END)
    input_field29.delete(0, tk.END)
    input_field30.delete(0, tk.END)
    input_field31.delete(0, tk.END)
    input_field32.delete(0, tk.END)
    input_field33.delete(0, tk.END)
    input_field34.delete(0, tk.END)
    input_field35.delete(0, tk.END)
    input_field36.delete(0, tk.END)
    input_field37.delete(0, tk.END)
    input_field38.delete(0, tk.END)
    input_field39.delete(0, tk.END)
    input_field40.delete(0, tk.END)
    input_field41.delete(0, tk.END)
    input_field42.delete(0, tk.END)
    # input_field43.delete(0, tk.END)
    # input_field44.delete(0, tk.END)
    input_field45.delete(0, tk.END)
    input_field46.delete(0, tk.END)
    input_field47.delete(0, tk.END)
    input_field48.delete(0, tk.END)
    input_field49.delete(0, tk.END)
    input_field50.delete(0, tk.END)
    input_field51.delete(0, tk.END)
    input_field52.delete(0, tk.END)
    input_field53.delete(0, tk.END)
    input_field54.delete(0, tk.END)
    input_field55.delete(0, tk.END)
    input_field56.delete(0, tk.END)
    messagebox.showinfo("Success", "Data pushed successfully!")

# Create a button widget to select the file and save the data
select_button = tk.Button(window, text="Push to Excel", command=save_data, bg="green", fg="white", font=("Helvetica", 12), width=10, bd=0)
select_button.grid(row=23, column=3)

# Run the tkinter mainloop
# window.mainloop()
window.withdraw()

# Start the login window
# login_window.mainloop()
window.mainloop()


























